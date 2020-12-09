from random import *
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "WooogySheet"

ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"])  # A1 셀의 정보 출력
print(ws["A1"].value)  # A1 셀의 값 출력
print(ws["A10"].value)  # 정보가 없을때는 None을 출력

# row = 1, 2, 3, ...
# column = A(1), B(2), C(3), ...
print(ws.cell(row=1, column=1).value)
print(ws.cell(row=1, column=2).value)  # B1의 값
print(ws.cell(column=2, row=1).value)  # Row, Column 바꿔도 됨

c = ws.cell(column=3, row=1, value=10)  # C1에 10 넣음
print(c.value)

# 반복문을 이용해서 랜덤 숫자 채우기
index = 1
for x in range(1, 11):  # row 10
    for y in range(1, 11):  # column 10
        # ws.cell(row=x, column=y, value=randint(0, 100)) # 0~100 사이 숫자
        ws.cell(row=x, column=y, value=index)
        index += 1

wb.save("sample.xlsx")
