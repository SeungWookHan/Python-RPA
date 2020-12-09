from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet()  # 새로운 시트 기본 이름으로 생성
ws.title = "Mysheet"  # Sheet 이름 변경
ws.sheet_properties.tabColor = "fff6ff"  # RGB 값을 넣어주면 탭 색상 변경

ws1 = wb.create_sheet("Yoursheet")  # 주어진 이름으로 시트 생성
ws2 = wb.create_sheet("Newsheet", 2)  # 2번째 인덱스에 시트 생성

new_ws = wb["Newsheet"]  # Dict 형태로 시트에 접근

print(wb.sheetnames)

# 시트 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("sample.xlsx")
