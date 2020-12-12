from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

# 번호, 영어, 수학
# 번호, (국어), 영어, 수학

# ws.move_range("B1:C11", rows=0, cols=1)  # 한열 옆으로 이동
# ws["B1"].value = "국어"  # B1셀에 국어 입력


# 번호, 영어, 수학
# 번호, 수학, 영어
# 5줄 아래로 내려서 왼쪽으로 한 열 이동 "수학" 컬럼을. 5줄 아래에 위치하는 것임!
ws.move_range("C1:C11", rows=5, cols=-1)

wb.save("sample_koream.xlsx")
