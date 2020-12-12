from openpyxl.utils.cell import coordinate_from_string
from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active

# 1줄씩 데이터 넣기
ws.append(["번호", "영어", "수학"])  # A B C
for i in range(1, 11):  # 10개 데이터 넣기
    ws.append([i, randint(0, 100), randint(0, 100)])

# col_B = ws["B"]  # 영어 column만 가져오기
# # print(col_B)
# # for cell in col_B:
# #     print(cell.value)

# # col_range = ws["B:C"]  # 영어, 수학 column 함께 가져오기
# # for cols in col_range:
# #     for cell in cols:
# #         print(cell.value)

# row_title = ws[1]  # 1번째 row만 가져오기
# # for cell in row_title:
# #     print(cell.value)


# row_range = ws[2:6]  # 2 번째 줄에서 6(포함) 기존 슬라이싱과 다름
# for rows in row_range:
#     for cell in rows:
#         # print(cell.value, end=" ")
#         # print(cell.coordinate, end=" ")  # A10,
#         xy = coordinate_from_string(cell.coordinate)
#         # print(xy, end=" ")  # 튜플 형태로 받아옴
#         print(xy[0], end="")  # A
#         print(xy[1], end=" ")  # B
#     print()

# # 전체 Rows
# print(tuple(ws.rows))

# # 전체 Columns
# print(tuple(ws.columns)) # 한 열씩 가져옴
# for row in tuple(ws.rows):
#     print(row[1].value)

# for column in tuple(ws.columns):
#     print(column[0].value)

# for row in ws.iter_rows():  # 전체 row에 대해
#     print(row[1].value)

# for column in ws.iter_cols():  # 전체 column에 대해
#     print(column[0].value)

# 1~5번째 줄까지, 2~3번째 열까지
# for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3):
#     # print(row[0].value, row[1].value)
#     print(row)

for column in ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3):
    print(column)

wb.save("sample.xlsx")
